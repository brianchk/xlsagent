"""Comments extractor (classic and threaded)."""

from __future__ import annotations

import re
from lxml import etree
from openpyxl.worksheet.worksheet import Worksheet

from ..models import CellReference, CommentInfo
from .base import BaseExtractor


class CommentExtractor(BaseExtractor):
    """Extracts comments (classic and threaded) from all sheets."""

    name = "comments"

    NAMESPACES = {
        "": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "tc": "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments",
    }

    def extract(self) -> list[CommentInfo]:
        """Extract all comments.

        Returns:
            List of CommentInfo objects
        """
        comments = []

        # Extract classic comments from openpyxl
        comments.extend(self._extract_classic_comments())

        # Extract threaded comments from XML
        comments.extend(self._extract_threaded_comments())

        return comments

    def _extract_classic_comments(self) -> list[CommentInfo]:
        """Extract classic comments via openpyxl."""
        comments = []

        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            if not isinstance(sheet, Worksheet):
                continue

            try:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.comment:
                            comments.append(CommentInfo(
                                location=CellReference(
                                    sheet=sheet_name,
                                    cell=cell.coordinate,
                                    row=cell.row,
                                    col=cell.column,
                                ),
                                author=cell.comment.author,
                                text=cell.comment.text or "",
                                is_threaded=False,
                            ))
            except Exception:
                continue

        return comments

    def _extract_threaded_comments(self) -> list[CommentInfo]:
        """Extract threaded comments from xl/threadedComments/."""
        comments = []

        contents = self.list_xlsx_contents()

        for item in contents:
            if "threadedComments" in item and item.endswith(".xml"):
                sheet_name = self._get_sheet_for_threaded_comments(item, contents)
                sheet_comments = self._parse_threaded_comments(item, sheet_name)
                comments.extend(sheet_comments)

        return comments

    def _get_sheet_for_threaded_comments(self, tc_path: str, contents: list[str]) -> str:
        """Determine which sheet threaded comments belong to."""
        # Parse the file number
        match = re.search(r"threadedComment(\d+)\.xml", tc_path)
        if not match:
            return "Unknown"

        try:
            tc_num = int(match.group(1))
            if 0 < tc_num <= len(self.workbook.sheetnames):
                return self.workbook.sheetnames[tc_num - 1]
        except Exception:
            pass

        return "Unknown"

    def _parse_threaded_comments(self, tc_path: str, sheet_name: str) -> list[CommentInfo]:
        """Parse threaded comments XML."""
        comments = []

        content = self.read_xml_from_xlsx(tc_path)
        if not content:
            return comments

        try:
            root = etree.fromstring(content)

            # Find all threadedComment elements
            # Group by reference cell for threading
            threads = {}

            for tc in root.findall(".//{http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments}threadedComment"):
                ref = tc.get("ref", "")
                parent_id = tc.get("parentId")

                text_elem = tc.find("{http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments}text")
                text = text_elem.text if text_elem is not None else ""

                # Get author from personId (would need to look up in xl/persons/person.xml)
                person_id = tc.get("personId")
                author = f"User {person_id}" if person_id else None

                comment = CommentInfo(
                    location=CellReference(
                        sheet=sheet_name,
                        cell=ref,
                        row=self._get_row_from_ref(ref),
                        col=self._get_col_from_ref(ref),
                    ),
                    author=author,
                    text=text or "",
                    is_threaded=True,
                )

                if parent_id:
                    # This is a reply
                    if ref not in threads:
                        threads[ref] = []
                    threads[ref].append(comment)
                else:
                    # This is a root comment
                    if ref not in threads:
                        threads[ref] = []
                    threads[ref].insert(0, comment)

            # Build final comment list with replies
            for ref, thread in threads.items():
                if thread:
                    root_comment = thread[0]
                    root_comment.replies = thread[1:]
                    comments.append(root_comment)

        except Exception:
            pass

        return comments

    def _get_row_from_ref(self, ref: str) -> int:
        """Extract row number from cell reference."""
        match = re.search(r"(\d+)", ref)
        return int(match.group(1)) if match else 0

    def _get_col_from_ref(self, ref: str) -> int:
        """Extract column number from cell reference."""
        match = re.search(r"([A-Z]+)", ref.upper())
        if not match:
            return 0

        col_str = match.group(1)
        col_num = 0
        for char in col_str:
            col_num = col_num * 26 + (ord(char) - ord("A") + 1)
        return col_num
