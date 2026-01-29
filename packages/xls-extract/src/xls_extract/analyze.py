"""
Main entry point for Excel workbook analysis.

This module provides the analyze() function which extracts structured data
from Excel workbooks.

Example:
    >>> from xls_extract import analyze
    >>> result = analyze("workbook.xlsx")
    >>> print(result.sheets)
"""

from __future__ import annotations

from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

import openpyxl

from .models import (
    WorkbookAnalysis,
    ExtractionError,
    ExtractionWarning,
)
from .extractors import (
    SheetExtractor,
    FormulaExtractor,
    NamedRangeExtractor,
    ConditionalFormatExtractor,
    DataValidationExtractor,
    PivotTableExtractor,
    ChartExtractor,
    TableExtractor,
    FilterExtractor,
    VBAExtractor,
    PowerQueryExtractor,
    ControlExtractor,
    ConnectionExtractor,
    CommentExtractor,
    HyperlinkExtractor,
    ProtectionExtractor,
    PrintSettingsExtractor,
    ErrorExtractor,
    DAXDetector,
)


@dataclass
class AnalysisOptions:
    """Options for controlling what gets extracted.

    Use this to skip expensive extraction steps if you don't need them,
    improving performance on large files.

    Attributes:
        extract_formulas: Extract all formulas (default: True).
        extract_vba: Extract VBA macros (default: True).
        extract_power_query: Extract Power Query M code (default: True).
        extract_charts: Extract chart information (default: True).
        extract_pivots: Extract pivot table info (default: True).
        extract_conditional_formats: Extract CF rules (default: True).
        extract_data_validations: Extract validation rules (default: True).
        extract_comments: Extract comments (default: True).
        extract_hyperlinks: Extract hyperlinks (default: True).
        extract_controls: Extract form controls (default: True).
        extract_connections: Extract data connections (default: True).
        extract_protection: Extract protection settings (default: True).
        extract_print_settings: Extract print settings (default: True).
        extract_errors: Detect error cells (default: True).
        include_formula_values: Include cached formula results (default: False).
        max_formulas: Maximum formulas to extract (default: None = unlimited).
        skip_sheets: List of sheet names to skip (default: empty).

    Example:
        >>> options = AnalysisOptions(
        ...     extract_vba=False,  # Skip VBA extraction
        ...     extract_power_query=False,  # Skip Power Query
        ...     max_formulas=1000,  # Limit formulas
        ... )
        >>> result = analyze("large_file.xlsx", options)
    """

    extract_formulas: bool = True
    extract_vba: bool = True
    extract_power_query: bool = True
    extract_charts: bool = True
    extract_pivots: bool = True
    extract_conditional_formats: bool = True
    extract_data_validations: bool = True
    extract_comments: bool = True
    extract_hyperlinks: bool = True
    extract_controls: bool = True
    extract_connections: bool = True
    extract_protection: bool = True
    extract_print_settings: bool = True
    extract_errors: bool = True
    include_formula_values: bool = False
    max_formulas: int | None = None
    skip_sheets: list[str] = field(default_factory=list)


def analyze(
    file_path: str | Path,
    options: AnalysisOptions | None = None,
) -> WorkbookAnalysis:
    """Analyze an Excel workbook and extract structured data.

    This is the main entry point for the library. It opens the Excel file,
    runs all configured extractors, and returns a WorkbookAnalysis object
    containing the results.

    Args:
        file_path: Path to the Excel file (.xlsx or .xlsm).
        options: Optional configuration for extraction. If not provided,
            all extraction features are enabled.

    Returns:
        WorkbookAnalysis containing all extracted data.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file is not a valid Excel file.

    Example:
        >>> result = analyze("financial_report.xlsx")
        >>> print(f"Found {len(result.formulas)} formulas")
        >>> for sheet in result.sheets:
        ...     print(f"  {sheet.name}: {sheet.row_count} rows")
    """
    if options is None:
        options = AnalysisOptions()

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    if path.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raise ValueError(f"Not a valid Excel file: {path}")

    # Initialize result
    result = WorkbookAnalysis(
        file_path=path,
        file_name=path.name,
        file_size=path.stat().st_size,
        is_macro_enabled=path.suffix.lower() in (".xlsm", ".xltm"),
    )

    errors: list[ExtractionError] = []
    warnings: list[ExtractionWarning] = []

    # Open workbook with openpyxl
    try:
        workbook = openpyxl.load_workbook(
            path,
            read_only=False,  # Need write access for some extractions
            data_only=False,  # Get formulas, not just values
            keep_vba=True,  # Preserve VBA for extraction
        )
    except Exception as e:
        raise ValueError(f"Could not open Excel file: {e}") from e

    try:
        # Run extractors
        _run_extractors(workbook, path, result, options, errors, warnings)
    finally:
        workbook.close()

    result.errors = errors
    result.warnings = warnings

    return result


def _run_extractors(
    workbook: openpyxl.Workbook,
    file_path: Path,
    result: WorkbookAnalysis,
    options: AnalysisOptions,
    errors: list[ExtractionError],
    warnings: list[ExtractionWarning],
) -> None:
    """Run all configured extractors."""

    # Always extract sheets first (needed by other extractors)
    try:
        extractor = SheetExtractor(workbook, file_path)
        result.sheets = extractor.extract()
    except Exception as e:
        errors.append(ExtractionError("sheets", str(e)))

    # Named ranges (needed for formula context)
    try:
        extractor = NamedRangeExtractor(workbook, file_path)
        result.named_ranges = extractor.extract()
    except Exception as e:
        errors.append(ExtractionError("named_ranges", str(e)))

    # Formulas
    if options.extract_formulas:
        try:
            extractor = FormulaExtractor(workbook, file_path)
            result.formulas = extractor.extract()
            if options.max_formulas and len(result.formulas) > options.max_formulas:
                result.formulas = result.formulas[: options.max_formulas]
                warnings.append(ExtractionWarning(
                    "formulas",
                    f"Limited to {options.max_formulas} formulas",
                ))
        except Exception as e:
            errors.append(ExtractionError("formulas", str(e)))

    # Conditional formatting
    if options.extract_conditional_formats:
        try:
            extractor = ConditionalFormatExtractor(workbook, file_path)
            result.conditional_formats = extractor.extract()
        except Exception as e:
            errors.append(ExtractionError("conditional_formats", str(e)))

    # Data validation
    if options.extract_data_validations:
        try:
            extractor = DataValidationExtractor(workbook, file_path)
            result.data_validations = extractor.extract()
        except Exception as e:
            errors.append(ExtractionError("data_validations", str(e)))

    # Pivot tables
    if options.extract_pivots:
        try:
            extractor = PivotTableExtractor(workbook, file_path)
            result.pivot_tables = extractor.extract()
        except Exception as e:
            errors.append(ExtractionError("pivot_tables", str(e)))

    # Charts
    if options.extract_charts:
        try:
            extractor = ChartExtractor(workbook, file_path)
            result.charts = extractor.extract()
        except Exception as e:
            errors.append(ExtractionError("charts", str(e)))

    # Tables
    try:
        extractor = TableExtractor(workbook, file_path)
        result.tables = extractor.extract()
    except Exception as e:
        errors.append(ExtractionError("tables", str(e)))

    # Filters
    try:
        extractor = FilterExtractor(workbook, file_path)
        result.auto_filters = extractor.extract()
    except Exception as e:
        errors.append(ExtractionError("filters", str(e)))

    # VBA
    if options.extract_vba and result.is_macro_enabled:
        try:
            extractor = VBAExtractor(workbook, file_path)
            vba_result = extractor.extract()
            result.vba_modules = vba_result.get("modules", [])
            result.vba_project_name = vba_result.get("project_name")
        except Exception as e:
            errors.append(ExtractionError("vba", str(e)))

    # Power Query
    if options.extract_power_query:
        try:
            extractor = PowerQueryExtractor(workbook, file_path)
            result.power_queries = extractor.extract()
        except Exception as e:
            errors.append(ExtractionError("power_query", str(e)))

    # Controls
    if options.extract_controls:
        try:
            extractor = ControlExtractor(workbook, file_path)
            result.controls = extractor.extract()
        except Exception as e:
            errors.append(ExtractionError("controls", str(e)))

    # Connections
    if options.extract_connections:
        try:
            extractor = ConnectionExtractor(workbook, file_path)
            result.connections = extractor.extract()
        except Exception as e:
            errors.append(ExtractionError("connections", str(e)))

    # Comments
    if options.extract_comments:
        try:
            extractor = CommentExtractor(workbook, file_path)
            result.comments = extractor.extract()
        except Exception as e:
            errors.append(ExtractionError("comments", str(e)))

    # Hyperlinks
    if options.extract_hyperlinks:
        try:
            extractor = HyperlinkExtractor(workbook, file_path)
            result.hyperlinks = extractor.extract()
        except Exception as e:
            errors.append(ExtractionError("hyperlinks", str(e)))

    # Protection
    if options.extract_protection:
        try:
            extractor = ProtectionExtractor(workbook, file_path)
            protection_result = extractor.extract()
            result.workbook_protection = protection_result.get("workbook")
            result.sheet_protections = protection_result.get("sheets", [])
        except Exception as e:
            errors.append(ExtractionError("protection", str(e)))

    # Print settings
    if options.extract_print_settings:
        try:
            extractor = PrintSettingsExtractor(workbook, file_path)
            result.print_settings = extractor.extract()
        except Exception as e:
            errors.append(ExtractionError("print_settings", str(e)))

    # Error cells
    if options.extract_errors:
        try:
            extractor = ErrorExtractor(workbook, file_path)
            error_result = extractor.extract()
            result.error_cells = error_result.get("errors", [])
            result.external_refs = error_result.get("external_refs", [])
        except Exception as e:
            errors.append(ExtractionError("errors", str(e)))

    # DAX detection
    try:
        detector = DAXDetector(workbook, file_path)
        dax_result = detector.extract()
        result.has_dax = dax_result.get("has_dax", False)
        result.dax_detection_note = dax_result.get("note")
    except Exception as e:
        errors.append(ExtractionError("dax_detection", str(e)))


@contextmanager
def open_workbook(file_path: str | Path) -> Iterator["WorkbookHandle"]:
    """Open a workbook for incremental extraction.

    Use this for very large files where you want to extract specific
    sheets without loading everything.

    Args:
        file_path: Path to the Excel file.

    Yields:
        WorkbookHandle for incremental extraction.

    Example:
        >>> with open_workbook("huge_file.xlsx") as wb:
        ...     print(wb.sheet_names)
        ...     data = wb.extract_sheet("Summary")
    """
    path = Path(file_path)
    workbook = openpyxl.load_workbook(
        path,
        read_only=True,
        data_only=False,
    )
    try:
        yield WorkbookHandle(workbook, path)
    finally:
        workbook.close()


class WorkbookHandle:
    """Handle for incremental workbook extraction.

    Provides access to workbook metadata and individual sheet extraction
    without loading everything into memory.
    """

    def __init__(self, workbook: openpyxl.Workbook, file_path: Path):
        self._workbook = workbook
        self._file_path = file_path

    @property
    def sheet_names(self) -> list[str]:
        """List of sheet names in the workbook."""
        return self._workbook.sheetnames

    @property
    def file_path(self) -> Path:
        """Path to the workbook file."""
        return self._file_path

    def extract_sheet(self, sheet_name: str) -> dict:
        """Extract data from a specific sheet.

        Args:
            sheet_name: Name of the sheet to extract.

        Returns:
            Dict containing sheet data.
        """
        if sheet_name not in self._workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        sheet = self._workbook[sheet_name]

        # Basic extraction
        return {
            "name": sheet_name,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
        }


def analyze_and_report(
    file_path: str | Path,
    output_dir: str | Path,
    options: AnalysisOptions | None = None,
    capture_screenshots: bool = True,
) -> WorkbookAnalysis:
    """Analyze an Excel workbook and generate complete reports.

    This is the high-level entry point that:
    1. Extracts all data from the workbook
    2. Captures screenshots (Windows only, requires xlwings)
    3. Generates HTML report
    4. Generates Markdown documentation

    Args:
        file_path: Path to the Excel file (.xlsx or .xlsm).
        output_dir: Directory to write reports and screenshots.
        options: Optional extraction configuration.
        capture_screenshots: Whether to capture screenshots (Windows only).

    Returns:
        WorkbookAnalysis containing all extracted data.

    Example:
        >>> from xls_extract import analyze_and_report
        >>> result = analyze_and_report("workbook.xlsx", "./analysis")
        >>> print(f"Report: {result.file_path.parent}/index.html")
    """
    import platform

    path = Path(file_path)
    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # Step 1: Extract data
    print(f"Analyzing: {path.name}", flush=True)
    result = analyze(path, options)

    # Step 2: Capture screenshots (Windows only)
    if capture_screenshots and platform.system() == "Windows":
        print("Capturing screenshots...", flush=True)
        try:
            from .screenshots import DesktopExcelScreenshotter
            screenshotter = DesktopExcelScreenshotter(out_path / "screenshots")
            result.screenshots = screenshotter.capture_all_sheets(path, result.sheets)
            print(f"  Captured {len(result.screenshots)} screenshots", flush=True)
        except ImportError:
            print("  Screenshots unavailable (install with: pip install xls-extract[screenshots])", flush=True)
        except Exception as e:
            print(f"  Screenshot capture failed: {e}", flush=True)
    elif capture_screenshots and platform.system() != "Windows":
        print("Screenshots only available on Windows", flush=True)

    # Step 3: Generate HTML report
    print("Generating HTML report...", flush=True)
    from .reports import HTMLReportBuilder
    html_builder = HTMLReportBuilder(result, out_path)
    html_path = html_builder.build()
    print(f"  Created: {html_path}", flush=True)

    # Step 4: Generate Markdown documentation
    print("Generating Markdown documentation...", flush=True)
    from .reports import MarkdownReportBuilder
    md_builder = MarkdownReportBuilder(result, out_path)
    md_builder.build()
    print(f"  Created: {out_path}/README.md", flush=True)

    return result
