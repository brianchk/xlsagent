"""Tests for report generators."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.models import WorkbookAnalysis
from src.extractors import SheetExtractor, FormulaExtractor, TableExtractor
from src.reports import HTMLReportBuilder, MarkdownReportBuilder


def create_analysis(workbook_path: Path) -> WorkbookAnalysis:
    """Helper to create an analysis from a workbook."""
    wb = load_workbook(workbook_path, data_only=False)

    analysis = WorkbookAnalysis(
        file_path=workbook_path,
        file_name=workbook_path.name,
        file_size=workbook_path.stat().st_size,
        is_macro_enabled=workbook_path.suffix.lower() in (".xlsm", ".xlsb"),
    )

    # Run extractors
    analysis.sheets = SheetExtractor(wb, workbook_path).extract()
    analysis.formulas = FormulaExtractor(wb, workbook_path).extract()
    analysis.tables = TableExtractor(wb, workbook_path).extract()

    wb.close()
    return analysis


class TestHTMLReportBuilder:
    """Tests for HTMLReportBuilder."""

    def test_generates_html_report(self, simple_workbook, temp_dir):
        analysis = create_analysis(simple_workbook)
        builder = HTMLReportBuilder(analysis, temp_dir)
        report_path = builder.build()

        assert report_path.exists()
        assert report_path.name == "report.html"

        content = report_path.read_text()
        assert "<html" in content
        assert analysis.file_name in content

    def test_html_contains_sheets_section(self, multi_sheet_workbook, temp_dir):
        analysis = create_analysis(multi_sheet_workbook)
        builder = HTMLReportBuilder(analysis, temp_dir)
        report_path = builder.build()

        content = report_path.read_text()
        assert "Visible" in content
        assert "Hidden" in content
        assert "VeryHidden" in content

    def test_html_contains_formulas_section(self, formula_workbook, temp_dir):
        analysis = create_analysis(formula_workbook)
        builder = HTMLReportBuilder(analysis, temp_dir)
        report_path = builder.build()

        content = report_path.read_text()
        assert "Formulas" in content
        assert "SUM" in content or "formula" in content.lower()

    def test_html_has_navigation(self, simple_workbook, temp_dir):
        analysis = create_analysis(simple_workbook)
        builder = HTMLReportBuilder(analysis, temp_dir)
        report_path = builder.build()

        content = report_path.read_text()
        assert "sidebar" in content
        assert "nav" in content.lower()

    def test_html_escapes_special_characters(self, temp_dir, simple_workbook):
        analysis = create_analysis(simple_workbook)
        # Inject potentially dangerous content
        analysis.file_name = "<script>alert('xss')</script>.xlsx"

        builder = HTMLReportBuilder(analysis, temp_dir)
        report_path = builder.build()

        content = report_path.read_text()
        # Should be escaped, not raw script tags
        assert "<script>alert" not in content
        assert "&lt;script&gt;" in content


class TestMarkdownReportBuilder:
    """Tests for MarkdownReportBuilder."""

    def test_generates_readme(self, simple_workbook, temp_dir):
        analysis = create_analysis(simple_workbook)
        builder = MarkdownReportBuilder(analysis, temp_dir)
        builder.build()

        readme = temp_dir / "README.md"
        assert readme.exists()

        content = readme.read_text()
        assert analysis.file_name in content
        assert "Summary" in content or "summary" in content

    def test_generates_summary(self, simple_workbook, temp_dir):
        analysis = create_analysis(simple_workbook)
        builder = MarkdownReportBuilder(analysis, temp_dir)
        builder.build()

        summary = temp_dir / "summary.md"
        assert summary.exists()

        content = summary.read_text()
        assert "sheet" in content.lower()

    def test_generates_sheet_files(self, multi_sheet_workbook, temp_dir):
        analysis = create_analysis(multi_sheet_workbook)
        builder = MarkdownReportBuilder(analysis, temp_dir)
        builder.build()

        sheets_dir = temp_dir / "sheets"
        assert sheets_dir.exists()

        index = sheets_dir / "_index.md"
        assert index.exists()

        # Check for individual sheet files
        sheet_files = list(sheets_dir.glob("*.md"))
        assert len(sheet_files) >= 2  # _index.md + at least one sheet

    def test_generates_formulas_index(self, formula_workbook, temp_dir):
        analysis = create_analysis(formula_workbook)
        builder = MarkdownReportBuilder(analysis, temp_dir)
        builder.build()

        formulas_dir = temp_dir / "formulas"
        assert formulas_dir.exists()

        index = formulas_dir / "_index.md"
        assert index.exists()

    def test_creates_features_directory(self, feature_workbook, temp_dir):
        analysis = create_analysis(feature_workbook)
        builder = MarkdownReportBuilder(analysis, temp_dir)
        builder.build()

        features_dir = temp_dir / "features"
        assert features_dir.exists()


class TestReportIntegration:
    """Integration tests for report generation."""

    def test_full_report_generation(self, feature_workbook, temp_dir):
        """Test complete report generation with all features."""
        wb = load_workbook(feature_workbook, data_only=False)

        # Create comprehensive analysis
        analysis = WorkbookAnalysis(
            file_path=feature_workbook,
            file_name=feature_workbook.name,
            file_size=feature_workbook.stat().st_size,
            is_macro_enabled=False,
        )

        # Import all extractors and run them
        from src.extractors import (
            SheetExtractor, FormulaExtractor, ConditionalFormatExtractor,
            DataValidationExtractor, FilterExtractor, CommentExtractor,
            HyperlinkExtractor,
        )

        analysis.sheets = SheetExtractor(wb, feature_workbook).extract()
        analysis.formulas = FormulaExtractor(wb, feature_workbook).extract()
        analysis.conditional_formats = ConditionalFormatExtractor(wb, feature_workbook).extract()
        analysis.data_validations = DataValidationExtractor(wb, feature_workbook).extract()
        analysis.auto_filters = FilterExtractor(wb, feature_workbook).extract()
        analysis.comments = CommentExtractor(wb, feature_workbook).extract()
        analysis.hyperlinks = HyperlinkExtractor(wb, feature_workbook).extract()

        wb.close()

        # Generate both reports
        html_builder = HTMLReportBuilder(analysis, temp_dir)
        html_path = html_builder.build()

        md_builder = MarkdownReportBuilder(analysis, temp_dir)
        md_builder.build()

        # Verify outputs
        assert html_path.exists()
        assert (temp_dir / "README.md").exists()
        assert (temp_dir / "summary.md").exists()
        assert (temp_dir / "sheets" / "_index.md").exists()
        assert (temp_dir / "formulas" / "_index.md").exists()
