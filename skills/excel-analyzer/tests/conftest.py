"""Pytest fixtures for Excel Analyzer tests."""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation


@pytest.fixture
def temp_dir():
    """Create a temporary directory for test outputs."""
    with tempfile.TemporaryDirectory() as td:
        yield Path(td)


@pytest.fixture
def simple_workbook(temp_dir) -> Path:
    """Create a simple workbook with basic data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Add some data
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Item 1"
    ws["B2"] = 100
    ws["A3"] = "Item 2"
    ws["B3"] = 200
    ws["A4"] = "Total"
    ws["B4"] = "=SUM(B2:B3)"

    path = temp_dir / "simple.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def formula_workbook(temp_dir) -> Path:
    """Create a workbook with various formula types."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulas"

    # Simple formulas
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=A1+A2"  # Simple

    # Lookup formulas
    ws["B1"] = "=VLOOKUP(A1,A1:B10,2,FALSE)"
    ws["B2"] = "=INDEX(A:A,MATCH(10,A:A,0))"

    # Aggregate formulas
    ws["C1"] = "=SUM(A1:A10)"
    ws["C2"] = "=SUMIF(A:A,\">10\")"
    ws["C3"] = "=COUNTIF(A:A,\">0\")"

    # Text formulas
    ws["D1"] = "=CONCATENATE(A1,B1)"
    ws["D2"] = "=LEFT(D1,5)"

    # Logical formulas
    ws["E1"] = "=IF(A1>10,\"Yes\",\"No\")"
    ws["E2"] = "=IFERROR(A1/0,\"Error\")"

    # Date formulas
    ws["F1"] = "=TODAY()"
    ws["F2"] = "=NOW()"

    # Dynamic array style (will have _xlfn prefix internally)
    ws["G1"] = "=_xlfn.XLOOKUP(A1,A:A,B:B)"
    ws["G2"] = "=_xlfn.FILTER(A1:A10,A1:A10>5)"
    ws["G3"] = "=_xlfn._xlpm.LAMBDA(x,x*2)(5)"

    # External reference
    ws["H1"] = "=[OtherBook.xlsx]Sheet1!A1"

    path = temp_dir / "formulas.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def multi_sheet_workbook(temp_dir) -> Path:
    """Create a workbook with multiple sheets including hidden ones."""
    wb = Workbook()

    # First sheet (visible)
    ws1 = wb.active
    ws1.title = "Visible"
    ws1["A1"] = "This is visible"

    # Second sheet (hidden)
    ws2 = wb.create_sheet("Hidden")
    ws2["A1"] = "This is hidden"
    ws2.sheet_state = "hidden"

    # Third sheet (very hidden)
    ws3 = wb.create_sheet("VeryHidden")
    ws3["A1"] = "This is very hidden"
    ws3.sheet_state = "veryHidden"

    # Fourth sheet with tab color
    ws4 = wb.create_sheet("Colored")
    ws4["A1"] = "This has a color"
    ws4.sheet_properties.tabColor = "FF0000"

    path = temp_dir / "multi_sheet.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def feature_workbook(temp_dir) -> Path:
    """Create a workbook with various Excel features."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Features"

    # Add data
    for i in range(1, 11):
        ws[f"A{i}"] = f"Item {i}"
        ws[f"B{i}"] = i * 10
        ws[f"C{i}"] = i * 100

    # Merged cells
    ws.merge_cells("D1:E2")
    ws["D1"] = "Merged"

    # Data validation (dropdown)
    dv = DataValidation(
        type="list",
        formula1='"Option1,Option2,Option3"',
        allow_blank=True
    )
    dv.add("F1:F10")
    ws.add_data_validation(dv)

    # Conditional formatting - color scale
    ws.conditional_formatting.add(
        "B1:B10",
        ColorScaleRule(
            start_type="min", start_color="FF0000",
            end_type="max", end_color="00FF00"
        )
    )

    # Conditional formatting - data bar
    ws.conditional_formatting.add(
        "C1:C10",
        DataBarRule(
            start_type="min", end_type="max",
            color="0000FF"
        )
    )

    # Conditional formatting - formula
    ws.conditional_formatting.add(
        "A1:A10",
        FormulaRule(
            formula=["$B1>50"],
            fill=None
        )
    )

    # Hyperlink
    ws["G1"] = "Click here"
    ws["G1"].hyperlink = "https://example.com"

    # Comment
    from openpyxl.comments import Comment
    ws["H1"].comment = Comment("This is a comment", "Author")

    # AutoFilter
    ws.auto_filter.ref = "A1:C10"

    path = temp_dir / "features.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def table_workbook(temp_dir) -> Path:
    """Create a workbook with structured tables."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tables"

    # Add table data
    headers = ["Name", "Department", "Salary"]
    data = [
        ["Alice", "Engineering", 100000],
        ["Bob", "Sales", 80000],
        ["Charlie", "Marketing", 75000],
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Create table
    from openpyxl.worksheet.table import Table, TableStyleInfo
    table = Table(displayName="Employees", ref="A1:C4")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    path = temp_dir / "tables.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def error_workbook(temp_dir) -> Path:
    """Create a workbook with error cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"

    # Create formulas that will result in errors
    ws["A1"] = "=1/0"  # #DIV/0!
    ws["A2"] = "=VLOOKUP(999,B:C,2,FALSE)"  # #N/A
    ws["A3"] = "=UnknownFunction()"  # #NAME?

    # Direct error values
    ws["B1"] = "#REF!"
    ws["B2"] = "#VALUE!"
    ws["B3"] = "#NULL!"

    path = temp_dir / "errors.xlsx"
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def named_range_workbook(temp_dir) -> Path:
    """Create a workbook with named ranges."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Add data
    ws["A1"] = 100
    ws["A2"] = 200
    ws["A3"] = 300

    # Create named ranges
    from openpyxl.workbook.defined_name import DefinedName

    # Global named range
    wb.defined_names.add(DefinedName("MyRange", attr_text="Data!$A$1:$A$3"))

    # Named constant
    wb.defined_names.add(DefinedName("TaxRate", attr_text="0.25"))

    # LAMBDA function (simulated - openpyxl stores as text)
    wb.defined_names.add(
        DefinedName("Double", attr_text="_xlfn._xlpm.LAMBDA(x,x*2)")
    )

    path = temp_dir / "named_ranges.xlsx"
    wb.save(path)
    wb.close()
    return path
