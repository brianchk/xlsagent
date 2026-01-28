"""Sheet metadata extractor."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from ..models import SheetInfo, SheetVisibility
from .base import BaseExtractor


class SheetExtractor(BaseExtractor):
    """Extracts sheet metadata including visibility, dimensions, and features."""

    name = "sheets"

    def extract(self) -> list[SheetInfo]:
        """Extract information about all sheets.

        Returns:
            List of SheetInfo objects
        """
        sheets = []

        for idx, sheet_name in enumerate(self.workbook.sheetnames):
            sheet = self.workbook[sheet_name]

            # Determine visibility
            visibility = self._get_visibility(sheet)

            # Get sheet dimensions
            used_range = None
            row_count = 0
            col_count = 0
            has_data = False

            if isinstance(sheet, Worksheet):
                if sheet.dimensions and sheet.dimensions != "A1:A1":
                    used_range = sheet.dimensions
                    try:
                        row_count = sheet.max_row or 0
                        col_count = sheet.max_column or 0
                        has_data = row_count > 0 and col_count > 0
                    except Exception:
                        pass

            # Check for various features
            sheet_info = SheetInfo(
                name=sheet_name,
                index=idx,
                visibility=visibility,
                used_range=used_range,
                row_count=row_count,
                col_count=col_count,
                has_data=has_data,
                has_formulas=self._has_formulas(sheet),
                has_charts=self._has_charts(sheet),
                has_pivots=self._has_pivots(sheet),
                has_tables=self._has_tables(sheet),
                has_comments=self._has_comments(sheet),
                has_conditional_formatting=self._has_conditional_formatting(sheet),
                has_data_validation=self._has_data_validation(sheet),
                has_hyperlinks=self._has_hyperlinks(sheet),
                has_merged_cells=self._has_merged_cells(sheet),
                merged_cell_ranges=self._get_merged_ranges(sheet),
                tab_color=self._get_tab_color(sheet),
            )

            sheets.append(sheet_info)

        return sheets

    def _get_visibility(self, sheet) -> SheetVisibility:
        """Get sheet visibility state."""
        try:
            state = sheet.sheet_state
            if state == "hidden":
                return SheetVisibility.HIDDEN
            elif state == "veryHidden":
                return SheetVisibility.VERY_HIDDEN
            else:
                return SheetVisibility.VISIBLE
        except Exception:
            return SheetVisibility.VISIBLE

    def _has_formulas(self, sheet) -> bool:
        """Check if sheet contains any formulas."""
        if not isinstance(sheet, Worksheet):
            return False
        try:
            for row in sheet.iter_rows(max_row=min(sheet.max_row or 0, 1000)):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        return True
        except Exception:
            pass
        return False

    def _has_charts(self, sheet) -> bool:
        """Check if sheet contains charts."""
        try:
            return len(sheet._charts) > 0
        except Exception:
            return False

    def _has_pivots(self, sheet) -> bool:
        """Check if sheet contains pivot tables."""
        try:
            return len(sheet._pivots) > 0
        except Exception:
            return False

    def _has_tables(self, sheet) -> bool:
        """Check if sheet contains structured tables."""
        try:
            return len(sheet.tables) > 0
        except Exception:
            return False

    def _has_comments(self, sheet) -> bool:
        """Check if sheet has any comments."""
        if not isinstance(sheet, Worksheet):
            return False
        try:
            # Check for legacy comments
            for row in sheet.iter_rows(max_row=min(sheet.max_row or 0, 500)):
                for cell in row:
                    if cell.comment:
                        return True
        except Exception:
            pass
        return False

    def _has_conditional_formatting(self, sheet) -> bool:
        """Check if sheet has conditional formatting."""
        try:
            return len(sheet.conditional_formatting) > 0
        except Exception:
            return False

    def _has_data_validation(self, sheet) -> bool:
        """Check if sheet has data validations."""
        try:
            return len(sheet.data_validations.dataValidation) > 0
        except Exception:
            return False

    def _has_hyperlinks(self, sheet) -> bool:
        """Check if sheet has hyperlinks."""
        try:
            return len(sheet.hyperlinks) > 0
        except Exception:
            return False

    def _has_merged_cells(self, sheet) -> bool:
        """Check if sheet has merged cells."""
        try:
            return len(sheet.merged_cells.ranges) > 0
        except Exception:
            return False

    def _get_merged_ranges(self, sheet) -> list[str]:
        """Get list of merged cell ranges."""
        try:
            return [str(r) for r in sheet.merged_cells.ranges]
        except Exception:
            return []

    def _get_tab_color(self, sheet) -> str | None:
        """Get sheet tab color if set."""
        try:
            if sheet.sheet_properties.tabColor:
                color = sheet.sheet_properties.tabColor
                if color.rgb:
                    return f"#{color.rgb}"
                elif color.theme is not None:
                    return f"theme:{color.theme}"
        except Exception:
            pass
        return None
